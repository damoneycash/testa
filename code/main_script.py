'''Projet d'aide à la lecture des tableaux Excel de l'INRAE'''
#obtenir la liste de scientifique
#split la data pour chaque scientifique
#scrap la data
#repartir sous forme de dictionnaire
#faire communiquer les dictionnaires
#ex : chercher toutes les 120 puis les trier par date puis par montant max
#resultat avec streamlit'''
#ou sur python directement..., avec input ou autre
from numpy import *
from pandas import *
from openpyxl import *
import datetime
import streamlit as st
import glob2

#st.set_page_config(layout="wide")
#GET LES EXCELS PRESENT DANS LE DOSSIER 
#print(option)

''' FEUILLE DATE '''


classeur = load_workbook('Cible.xlsx')
wb = classeur.sheetnames
wa = wb[0]
drap1 = classeur[wa]
#print(drap1)
#Get date'''
e = str(datetime.datetime.today())
e1 = e.split(" ")[0]
e2 = e1.split("-")[0]
year = int(e2)+1

L = []
for i in drap1["F"]:
    L.append(i.value)
L.pop(0)
L.pop(0)
L.pop(0)
#print(L)
P = []
for i in range(len(L)):
    if str(L[i]) == 'None':
        L[i] = 1
    elif str(L[i]) == 'Terminé':
        L[i] = datetime.datetime(2020, 11, 11, 0, 0)
    elif str(L[i]) == 'PROLONGE':
        L[i] = 1
    elif str(L[i]) == ' CHAQUE ANNEE':
        L[i] = datetime.datetime(year, 11, 11, 0, 0)
W = []
z = 0
for i in range(len(L)):
    if L[i] == 1:
        z += 1
    else :
         W.append(L[i])
Ldate = array(W)
#print(Ld)

''' FILE DATA '''


#file = glob2.glob('EXCEL\*.xlsx')
classeur = load_workbook('Cible.xlsx') #'EXCEL\ps(168).xlsx'
wb = classeur.sheetnames
wa = wb[1]
drap = classeur[wa]
#print(drap)
#print(wb[1])
#print(classeur.get_sheet_names())
#print(drap["H"])

#Get liste Scientifiques'''
L = []
for i in drap["H"]:
    L.append(i.value)
L.pop(0)
L.sort()
Sc = array(L).unique()
#print(Sc)

#Get liste lignes analytiques
L = []
for i in drap["L"]:
    L.append(i.value)
L.pop(0)
L.sort()
La = array(L).unique()
'''
#Get liste intitulé
L = []
for i in drap["I"]:
    L.append(i.value)
L.pop(0)
L.sort()
Li = array(L).unique()
#print(Li)
'''

'''
#GET LISTE DATES
G = []
for i in drap["R"]:
    G.append(i.value)
G.pop(0)
Lg = array(G).unique()

L = []
for i in drap["R"]:
    L.append(i.value)
L.pop(0)
M = []
for i in L:
    a = str(i)
    split1 = a.split(" ")[0]
    M.append(split1)
M.sort()
#print(len(M))
cestuntest = M
Ld = array(M).unique()
#print(Ld)
'''
#Get tout e la data
a = ''
for line in drap:
    for i in line:
        #print(str(i.value))
        a =  a + "{" + str(i.value)
#print(a)
#Split la data
Sc_data = []
Sc_data_final = []
Sc_data.append(a.split("BPP"))
V = Sc_data[0][0]
Sc_data[0].pop(0)
#print(Sc_data[0][1])
for j in range(len(Sc_data[0])-1):
    A = []
    a = ''
    a = str(Sc_data[0][j])
    split1 = a.split('{')
    #print(split1)
    for i in split1:
        A.append(i)
    A.pop(0)
    A.pop(len(split1)-2)
    #print(A)
    Sc_data_final.append(A)
a = ''
a = str(Sc_data[0][len(Sc_data)-2])
split1 = a.split('{')
#print(str(Sc_data[0][len(Sc_data)-2]))
#print(len(Sc_data[0]))
A = []
for i in split1:
    A.append(i)
A.pop(0)
#print(A)
Sc_data_final.append(A)
#for k in range(len(Sc_data_final)):
    #print(Sc_data_final[k])

#Répartir la data pour chaque scientifique
K = []
L = []
Z = []
for i in range(len(Sc)):
    K.append(L)
#print(K)
J = []
w = 0
for i in range(len(Sc)): #len(Sc)
    L = []
    e = 0
    for k in range(len(Sc_data_final)):
        if str(Sc_data_final[k][6]) == str(Sc[i]):
            #K[i].append(Sc_data_final[k])
            L.append(e)
            #print(Sc_data_final[e][6])   #pass      
        e += 1
    for j in L:
        #print(j)
        J.append(str(Sc_data_final[int(j)]))
    Z.append(J)
    L = []
    J = []
#print(Z[7:20])
L1, L2, L3 = [], [], []
for i in range(len(Z)): 
    L2 = []
    for j in Z[i]:
        L1 = []
        #print(j)
        r = ''
        split2 = j.split("', '")
        for z in range(2, len(split2[0])):
            r += split2[0][z]
        split2[0] = r    
        r = ''
        for z in range(len(split2[len(split2)-1])-2):
            r += split2[len(split2)-1][z]
        split2[len(split2)-1] = r
        #print(split2)
        for k in split2:
            #print(k)
            #print(k)
            L1.append(k)
        L2.append(L1)
    L3.append(L2)
#print(L3[0][0][0])
Sc_data = L3


#Get liste intitulé
L = []
for i in range(len(Sc_data)):
    for j in range(len(Sc_data[i])):
        L.append(Sc_data[i][j][7])
Li = array(L).unique()
#print(Li)


#print(Sc_data)
''' JONCTION FEUILLE '''
key = []
for i in range(len(Li)):
    B = []
    B.append(Ldate[i])
    B.append(Li[i])
    key.append(B)
#print(key)

''' INSERTION JONCTION DATA'''
G = []
for i in range(len(Sc_data)):
    for j in range(len(Sc_data[i])):
        for k in range(len(key)):
            if key[k][1] == Sc_data[i][j][7]:
                Sc_data[i][j].append(key[k][0])
                G.append(key[k][0])
for i in range(len(Sc_data)):
    for j in range(len(Sc_data[i])):
        #Sc_data[i][j] = Sc_data[i][j][:16] + Sc_data[i][j][17:]
        if len(Sc_data[i][j]) != len(Sc_data[0][0]):
            Sc_data[i][j].append(datetime.datetime(999, 1, 1, 0, 0))
            G.append(datetime.datetime(999, 1, 1, 0, 0))

#print(Sc_data)
L2 = []
for i in range(len(Sc_data)):
    L = []
    for j in range(len(Sc_data[i])):
        L.append(Sc_data[i][j][16])
    M = []
    for r in L:
        a = str(r)
        split1 = a.split(" ")[0]
        M.append(split1)
        #print(M)
    D = M.copy()
    M.sort()
    L1 = []
        #print(D)
        #print(M)
    for z in range(len(M)):
        for l in range(len(D)):
            if M[z] == D[l]:
                Sc_data[i][l][16] = M[z]
                L1.append(Sc_data[i][l])
                
    #print(L1)        
    L2.append(L1)
Sc_data = L2
#print(Sc_data)


#MISE EN FORME DES EN-TETES
Headers = []
V = str(V) + "Date"
b = str(V)
split = b.split("{")
for i in split:
    Headers.append(i)
Headers.pop(0)
Headers.pop(0)
#Headers.pop(len(Headers)-1)
#print(Headers)

Headers_d = {}
z = 0
for i in Headers:
    Headers_d[z] = i
    z += 1
'''
G = []
for i in range(len(Sc_data)):
    for j in range(len(Sc_data[i])):
        G.append(Sc_data[i][j][16])
print(G)
'''

def date_expire():
    Urg = []
    Warn = []
    K = []
    g = st.slider(label = "Choisissez la date limite de date de fin en jours à partir d'aujourd'hui", min_value= 1, max_value=120)
    e = datetime.datetime.today()
    for i in range(len(G)):
        a = str(G[i])
        split1 = a.split(" ")[0]
        K.append(split1)
        if int((G[i]-e).days) >= 0:
            if int((G[i]-e).days) <= g:
                Urg.append(K[i]) #M
    #print(Urg)
    for j in range(len(Sc_data)):        
        for l in range(len(Sc_data[j])):
            #print(Sc_data[j][l][16])
            for z in range(len(Urg)):
                #print(Sc_data[j][l][16])
                if str(Sc_data[j][l][16]) == str(Urg[z]):
                    Warn.append(Sc_data[j][l])
            
    #print(Warn[1][16])
    L2 = []
    M = []
    for i in range(len(Warn)):
        M.append(Warn[i][16])
        #print(M)
    M.sort()
    L1 = []
        #print(D)
        #print(M)
    for z in range(len(M)):
        for j in range(len(Warn)):
            if M[z] == Warn[j][16]:
                L1.append(Warn[j])

    #print(M)   
    #print(L1)
    L2.append(L1)
    Warn = L1

    if not Warn:
        return None

    df = DataFrame(Warn)
    df.columns=Headers
    st.dataframe(df, use_container_width= True, height= 700, hide_index= True)
    st.column_config.TextColumn(width="small")
    return
#print(Sc_data)


#METTRE EN FORME AVEC STREAMLIT SCIENTIFIQUE
def scientifique():
    #st.header(":blue[INRAE BPP]")
    Sc_dico = {}
    z = 0
    for i in Sc:
        Sc_dico[i] = z
        z += 1
    option = st.selectbox("CHOISISSEZ LE SCIENTIFIQUE ⬇️",Sc)
    m = Sc_dico[option]
    st.title(Sc[m])
    #print(Sc_data)
    df = DataFrame(Sc_data[m])
    df.columns=Headers
    st.dataframe(df, use_container_width= True, height= 400, hide_index= True)
    st.column_config.TextColumn(width="small")
    return
#scientifique()
#print(df)
#print(len(Sc_data))

#REPARTIR LA DATA POUR CHAQUE LIGNE ANALYTIQUE
K = []
L = []
Z = []
for i in range(len(La)):
    K.append(L)
#print(K)
J = []
w = 0
#print(Sc_data_final)
for i in range(len(La)): 
    L = []
    e = 0
    for k in range(len(Sc_data_final)):
        if str(Sc_data_final[k][10]) == str(La[i]):
            #K[i].append(Sc_data_final[k])
            L.append(e)
            #print(Sc_data_final[e][6])   #pass      
        e += 1
    for j in L:
        #print(j)
        J.append(str(Sc_data_final[int(j)]))
    Z.append(J)
    L = []
    J = []
#print(Z[1])
L1, L2, L3 = [], [], []
for i in range(len(Z)): 
    L2 = []
    for j in Z[i]:
        L1 = []
        #print(j)
        r = ''
        split2 = j.split("', '")
        for z in range(2, len(split2[0])):
            r += split2[0][z]
        split2[0] = r    
        r = ''
        for z in range(len(split2[len(split2)-1])-2):
            r += split2[len(split2)-1][z]
        split2[len(split2)-1] = r
        #print(split2)
        for k in split2:
            #print(k)
            #print(k)
            L1.append(k)
        L2.append(L1)
    L3.append(L2)
#print(L3[0][0][0])


La_data = L3
#print(L3[3])
''' INSERTION JONCTION DATA'''
G = []
for i in range(len(La_data)):
    for j in range(len(La_data[i])):
        for k in range(len(key)):
            if key[k][1] == La_data[i][j][7]:
                La_data[i][j].append(key[k][0])
                G.append(key[k][0])
for i in range(len(La_data)):
    for j in range(len(La_data[i])):
        #Sc_data[i][j] = Sc_data[i][j][:16] + Sc_data[i][j][17:]
        if len(La_data[i][j]) != len(La_data[0][0]):
            La_data[i][j].append(datetime.datetime(999, 1, 1, 0, 0))
            G.append(datetime.datetime(999, 1, 1, 0, 0))

#print(Sc_data)
L2 = []
for i in range(len(La_data)):
    L = []
    for j in range(len(La_data[i])):
        L.append(La_data[i][j][16])
    M = []
    for r in L:
        a = str(r)
        split1 = a.split(" ")[0]
        M.append(split1)
        #print(M)
    D = M.copy()
    M.sort()
    L1 = []
        #print(D)
        #print(M)
    for z in range(len(M)):
        for l in range(len(D)):
            if M[z] == D[l]:
                La_data[i][l][16] = M[z]
                L1.append(La_data[i][l])
                
    #print(L1)        
    L2.append(L1)
La_data = L2
#print(La_data)

#METTRE EN FORME AVEC STREAMLIT ANALYTIQUE
def analytique():
    #st.header(":blue[INRAE BPP]")
    La_dico = {}
    z = 0
    for i in La:
        La_dico[i] = z
        z += 1
    option = st.selectbox("CHOISISSEZ LA LIGNE ANALYTIQUE ⬇️",La)
    m = La_dico[option]
    a = "Code :  " + str(La[m]) 
    st.title(a)
    df = DataFrame(La_data[m])
    df.columns=Headers
    st.dataframe(df, use_container_width= True, height=700, hide_index=True)
    st.column_config.TextColumn(width="small")
    return
#analytique()
#print(df)
#print(len(Sc_data))


''' JONCTION FILE'''














#GET LES CHIFFRES CLES

budg = 0
for i in range(len(Sc_data_final)):
    budg = budg + float(Sc_data_final[i][12])
budget = round(budg, 2)

dep = 0
for i in range(len(Sc_data_final)):
    dep = dep + float(Sc_data_final[i][13])
depense = round(dep, 2)


budg_d = 0
#print(Sc_data_final[len(Sc_data_final)-2][15])
for i in range(len(Sc_data_final)):
    #print(Sc_data)
    budg_d = budg_d + float(Sc_data_final[i][15])
budget_disponible = round(budg_d, 2)

nombre_mission = len(Sc_data_final)+1


Table_final = []
for i in range(len(Sc_data)):
    for j in range(len(Sc_data[i])):
        Table_final.append(Sc_data[i][j])




#METTRE L ENSEMBLE DU DATAFRAME
def table_complete():
    df = DataFrame(Table_final)
    df.columns=Headers
    #st.markdown(' ')
    st.dataframe(data=df, use_container_width= True, height=700)
    st.column_config.TextColumn(width="small")
    return 
#table_complete()

#FONCTIONS PRATIQUE STREMLIT
def saut_ligne(i):
    for i in range(int(i)):
        st.write("")
    return













def scientifique_drap(drap):
    L = []
    for i in drap["H"]:
        L.append(i.value)
    L.pop(0)
    L.sort()
    Sc = array(L).unique()


#Get toute la data
    a = ''
    for line in drap:
        for i in line:
        #print(str(i.value))
            a =  a + "{" + str(i.value)

#Split la data
    Sc_data = []
    Sc_data_final = []
    Sc_data.append(a.split("BPP"))
    V = Sc_data[0][0]
    Sc_data[0].pop(0)
#print(Sc_data[0][1])
    for j in range(len(Sc_data[0])-1):
        A = []
        a = ''
        a = str(Sc_data[0][j])
        split1 = a.split('{')
    #print(split1)
        for i in split1:
            A.append(i)
        A.pop(0)
        A.pop(len(split1)-2)
    #print(A)
        Sc_data_final.append(A)
    a = ''
    a = str(Sc_data[0][len(Sc_data)-2])
    split1 = a.split('{')
#print(str(Sc_data[0][len(Sc_data)-2]))
#print(len(Sc_data[0]))
    A = []
    for i in split1:
        A.append(i)
    A.pop(0)
#print(A)
    Sc_data_final.append(A)
#print(Sc_data_final[100:400])
#for k in range(len(Sc_data_final)):
    #print(Sc_data_final[k])

#Répartir la data pour chaque scientifique
    K = []
    L = []
    Z = []  
    for i in range(len(Sc)):
        K.append(L)
#print(K)
    J = []
    w = 0
    for i in range(len(Sc)): #len(Sc)
        L = []
        e = 0
        for k in range(len(Sc_data_final)):
            if str(Sc_data_final[k][6]) == str(Sc[i]):
            #K[i].append(Sc_data_final[k])
                L.append(e)
            #print(Sc_data_final[e][6])   #pass      
            e += 1
        for j in L:
        #print(j)
            J.append(str(Sc_data_final[int(j)]))
        Z.append(J)
        L = []
        J = []
#print(Z[7:20])
    L1, L2, L3 = [], [], []
    for i in range(len(Z)): 
        L2 = []
        for j in Z[i]:
            L1 = []
        #print(j)
            r = ''
            split2 = j.split("', '")
            for z in range(2, len(split2[0])):
                r += split2[0][z]
            split2[0] = r    
            r = ''
            for z in range(len(split2[len(split2)-1])-2):
                r += split2[len(split2)-1][z]
            split2[len(split2)-1] = r
        #print(split2)
            for k in split2:
            #print(k)
            #print(k)
                L1.append(k)
            L2.append(L1)
        L3.append(L2)
#print(L3[0][0][0])
    Sc_data = L3
    L2 = []
    for i in range(len(Sc_data)):
        L = []
        for j in range(len(Sc_data[i])):
            L.append(Sc_data[i][j][16])
        M = []
        for r in L:
            a = str(r)
            split1 = a.split(" ")[0]
            M.append(split1)
        #print(M)
        D = M.copy()
        M.sort()
        L1 = []
        #print(D)
        #print(M)
        for z in range(len(M)):
            for l in range(len(D)):
                if M[z] == D[l]:
                    Sc_data[i][l][16] = M[z]
                    L1.append(Sc_data[i][l])
    #print(L1)        
        L2.append(L1)
    Sc_data = L2

#MISE EN FORME DES EN-TETES
    Headers = []
    b = str(V)
    split = b.split("{")
    for i in split:
        Headers.append(i)
    Headers.pop(0)
    Headers.pop(0)
    Headers.pop(len(Headers)-1)
#print(Headers)

    Headers_d = {}
    z = 0
    for i in Headers:
        Headers_d[z] = i
        z += 1

    Sc_dico = {}
    z = 0
    for i in Sc:
        Sc_dico[i] = z
        z += 1
    option = st.selectbox("CHOISISSEZ LE SCIENTIFIQUE ⬇️",Sc)
    m = Sc_dico[option]
    st.title(Sc[m])
    df = DataFrame(Sc_data[m])
    df.columns=Headers
    st.dataframe(df, use_container_width= True, height= 400, hide_index= True)
    st.column_config.TextColumn(width="small")
    
    return


def analytique_drap(drap):
    L = []
    for i in drap["H"]:
        L.append(i.value)
    L.pop(0)
    L.sort()
    Sc = array(L).unique()

    L = []
    for i in drap["L"]:
        L.append(i.value)
    L.pop(0)
    L.sort()
    La = array(L).unique()

#Get toute la data
    a = ''
    for line in drap:
        for i in line:
        #print(str(i.value))
            a =  a + "{" + str(i.value)

#Split la data
    Sc_data = []
    Sc_data_final = []
    Sc_data.append(a.split("BPP"))
    V = Sc_data[0][0]
    Sc_data[0].pop(0)
#print(Sc_data[0][1])
    for j in range(len(Sc_data[0])-1):
        A = []
        a = ''
        a = str(Sc_data[0][j])
        split1 = a.split('{')
    #print(split1)
        for i in split1:
            A.append(i)
        A.pop(0)
        A.pop(len(split1)-2)
    #print(A)
        Sc_data_final.append(A)
    a = ''
    a = str(Sc_data[0][len(Sc_data)-2])
    split1 = a.split('{')
#print(str(Sc_data[0][len(Sc_data)-2]))
#print(len(Sc_data[0]))
    A = []
    for i in split1:
        A.append(i)
    A.pop(0)
#print(A)
    Sc_data_final.append(A)
#print(Sc_data_final[100:400])
#for k in range(len(Sc_data_final)):
    #print(Sc_data_final[k])

#Répartir la data pour chaque scientifique
    
#MISE EN FORME DES EN-TETES
    Headers = []
    b = str(V)
    split = b.split("{")
    for i in split:
        Headers.append(i)
    Headers.pop(0)
    Headers.pop(0)
    Headers.pop(len(Headers)-1)
#print(Headers)

    Headers_d = {}
    z = 0
    for i in Headers:
        Headers_d[z] = i
        z += 1
    K = []
    L = []
    Z = []
    for i in range(len(La)):
        K.append(L)
#print(K)
    J = []
    w = 0
#print(Sc_data_final)
    for i in range(len(La)): 
        L = []
        e = 0
        for k in range(len(Sc_data_final)):
            if str(Sc_data_final[k][10]) == str(La[i]):
            #K[i].append(Sc_data_final[k])
                L.append(e)
            #print(Sc_data_final[e][6])   #pass      
            e += 1
        for j in L:
        #print(j)
            J.append(str(Sc_data_final[int(j)]))
        Z.append(J)
        L = []
        J = []
#print(Z[1])
    L1, L2, L3 = [], [], []
    for i in range(len(Z)): 
        L2 = []
        for j in Z[i]:
            L1 = []
        #print(j)
            r = ''
            split2 = j.split("', '")
            for z in range(2, len(split2[0])):
                r += split2[0][z]
            split2[0] = r    
            r = ''
            for z in range(len(split2[len(split2)-1])-2):
                r += split2[len(split2)-1][z]
            split2[len(split2)-1] = r
        #print(split2)
            for k in split2:
            #print(k)
            #print(k)
                L1.append(k)
            L2.append(L1)
        L3.append(L2)
#print(L3[0][0][0])

    Sc_data = L3
    L2 = []
    for i in range(len(Sc_data)):
        L = []
        for j in range(len(Sc_data[i])):
            L.append(Sc_data[i][j][16])
        M = []
        for r in L:
            a = str(r)
            split1 = a.split(" ")[0]
            M.append(split1)
        #print(M)
        D = M.copy()
        M.sort()
        L1 = []
        #print(D)
        #print(M)
        for z in range(len(M)):
            for l in range(len(D)):
                if M[z] == D[l]:
                    Sc_data[i][l][16] = M[z]
                    L1.append(Sc_data[i][l])
    #print(L1)        
        L2.append(L1)
    Sc_data = L2
#print(L3[3])

#METTRE EN FORME AVEC STREAMLIT ANALYTIQUE
    #st.header(":blue[INRAE BPP]")
    La_dico = {}
    z = 0
    for i in La:
        La_dico[i] = z
        z += 1
    option = st.selectbox("CHOISISSEZ LA LIGNE ANALYTIQUE ⬇️",La)
    m = La_dico[option]
    a = "Code :  " + str(La[m]) 
    st.title(a)
    df = DataFrame(Sc_data[m])
    df.columns=Headers
    st.dataframe(df, use_container_width= True, height=700, hide_index=True)
    st.column_config.TextColumn(width="small")
    
    return


def table_drap(drap):
    L = []
    for i in drap["H"]:
        L.append(i.value)
    L.pop(0)
    L.sort()
    Sc = array(L).unique()


#Get toute la data
    a = ''
    for line in drap:
        for i in line:
        #print(str(i.value))
            a =  a + "{" + str(i.value)

#Split la data
    Sc_data = []
    Sc_data_final = []
    Sc_data.append(a.split("BPP"))
    V = Sc_data[0][0]
    Sc_data[0].pop(0)
#print(Sc_data[0][1])
    for j in range(len(Sc_data[0])-1):
        A = []
        a = ''
        a = str(Sc_data[0][j])
        split1 = a.split('{')
    #print(split1)
        for i in split1:
            A.append(i)
        A.pop(0)
        A.pop(len(split1)-2)
    #print(A)
        Sc_data_final.append(A)
    a = ''
    a = str(Sc_data[0][len(Sc_data)-2])
    split1 = a.split('{')
#print(str(Sc_data[0][len(Sc_data)-2]))
#print(len(Sc_data[0]))
    A = []
    for i in split1:
        A.append(i)
    A.pop(0)
#print(A)
    Sc_data_final.append(A)
    for i in range(len(Sc_data_final)):
        Sc_data_final[i][16] = Ldate[i]
    Headers = []
    b = str(V)
    split = b.split("{")
    for i in split:
        Headers.append(i)
    Headers.pop(0)
    Headers.pop(0)
    Headers.pop(len(Headers)-1)
#print(Headers)
    Headers_d = {}
    z = 0
    for i in Headers:
        Headers_d[z] = i
        z += 1
    df = DataFrame(Sc_data_final)
    df.columns=Headers
    #st.markdown(' ')
    st.dataframe(data=df, use_container_width= True, height=700)
    st.column_config.TextColumn(width="small")
    return
